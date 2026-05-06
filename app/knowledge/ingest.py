"""
Helpers for ingesting policy Q&A files.
"""

from __future__ import annotations

import csv
import logging
import re
import zipfile
from dataclasses import dataclass
from pathlib import Path
from typing import Any
from xml.etree import ElementTree as ET

QUESTION_HEADERS = {
    "question",
    "questions",
    "query",
    "faq_question",
    "user_question",
}

ANSWER_HEADERS = {
    "answer",
    "answers",
    "response",
    "faq_answer",
    "bot_answer",
}

logger = logging.getLogger(__name__)


@dataclass
class PolicyDocChunk:
    chunk_index: int
    content: str
    section_title: str | None = None
    page_number: int | None = None
    char_start: int = 0
    char_end: int = 0
    metadata: dict[str, Any] | None = None


def _normalize_header(name: str) -> str:
    return "".join(ch.lower() if ch.isalnum() else "_" for ch in name).strip("_")


def _match_header(headers: list[str], candidates: set[str]) -> int:
    normalized = [_normalize_header(h) for h in headers]
    for idx, header in enumerate(normalized):
        if header in candidates:
            return idx
    raise RuntimeError(
        f"Could not find required header in file. Expected one of: {sorted(candidates)}"
    )


def read_policy_rows(file_path: Path) -> list[dict[str, Any]]:
    suffix = file_path.suffix.lower()
    if suffix == ".csv":
        return _read_csv(file_path)
    if suffix in {".xlsx", ".xlsm"}:
        return _read_xlsx(file_path)
    raise ValueError(f"Unsupported file extension: {suffix}")


def read_policy_docx_chunks(
    file_path: Path,
    *,
    chunk_size_chars: int = 2200,
    chunk_overlap_chars: int = 300,
    min_chunk_chars: int = 220,
) -> list[PolicyDocChunk]:
    """
    Parse DOCX and return heading-aware chunks for long-form policy retrieval.
    """
    if file_path.suffix.lower() != ".docx":
        raise ValueError("DOCX ingestion requires a .docx file")

    try:
        from docx import Document
    except ImportError as exc:
        raise RuntimeError("python-docx is required for .docx ingestion. Install: pip install python-docx") from exc

    try:
        sections = _extract_docx_sections_via_python_docx(file_path)
    except Exception as exc:
        logger.warning("python-docx parse failed for %s: %s. Falling back to XML parser.", file_path, str(exc))
        sections = _extract_docx_sections_via_xml(file_path)

    chunks: list[PolicyDocChunk] = []
    chunk_index = 0
    for heading, section_text in sections:
        chunks.extend(
            _chunk_section_text(
                section_text,
                section_title=heading,
                start_index=chunk_index,
                chunk_size_chars=chunk_size_chars,
                chunk_overlap_chars=chunk_overlap_chars,
                min_chunk_chars=min_chunk_chars,
            )
        )
        chunk_index = len(chunks)

    return chunks


def _read_csv(file_path: Path) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    with file_path.open("r", newline="", encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        headers = next(reader, None)
        if not headers:
            return rows
        q_idx = _match_header([str(h) for h in headers], QUESTION_HEADERS)
        a_idx = _match_header([str(h) for h in headers], ANSWER_HEADERS)
        for i, row in enumerate(reader, start=2):
            question = row[q_idx].strip() if q_idx < len(row) and row[q_idx] is not None else ""
            answer = row[a_idx].strip() if a_idx < len(row) and row[a_idx] is not None else ""
            rows.append({"question": question, "answer": answer, "row_number": i})
    return rows


def _read_xlsx(file_path: Path) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required for xlsx ingestion. Install: pip install openpyxl") from exc

    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active
    headers = [str(c.value).strip() if c.value is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
    q_idx = _match_header(headers, QUESTION_HEADERS)
    a_idx = _match_header(headers, ANSWER_HEADERS)

    rows: list[dict[str, Any]] = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        question = "" if row[q_idx] is None else str(row[q_idx]).strip()
        answer = "" if row[a_idx] is None else str(row[a_idx]).strip()
        rows.append({"question": question, "answer": answer, "row_number": row_idx})
    wb.close()
    return rows


def _clean_text(text: str) -> str:
    return re.sub(r"\s+", " ", text).strip()


def _extract_docx_sections_via_python_docx(file_path: Path) -> list[tuple[str, str]]:
    from docx import Document

    doc = Document(str(file_path))
    section_title = "Policy Document"
    sections: list[tuple[str, str]] = []
    current_lines: list[str] = []

    for paragraph in doc.paragraphs:
        raw = paragraph.text or ""
        text = _clean_text(raw)
        if not text:
            continue

        style_name = (paragraph.style.name or "").lower() if paragraph.style else ""
        if "heading" in style_name:
            if current_lines:
                sections.append((section_title, "\n".join(current_lines).strip()))
                current_lines = []
            section_title = text
            continue

        current_lines.append(text)

    if current_lines:
        sections.append((section_title, "\n".join(current_lines).strip()))
    return sections


def _extract_docx_sections_via_xml(file_path: Path) -> list[tuple[str, str]]:
    """
    Fallback parser for malformed .docx archives where python-docx fails
    (e.g., broken rel targets like word/NULL).
    """
    ns = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}

    with zipfile.ZipFile(file_path, "r") as zf:
        if "word/document.xml" not in zf.namelist():
            raise ValueError("Invalid DOCX: missing word/document.xml")
        doc_root = ET.fromstring(zf.read("word/document.xml"))

        style_to_name: dict[str, str] = {}
        if "word/styles.xml" in zf.namelist():
            try:
                styles_root = ET.fromstring(zf.read("word/styles.xml"))
                for style in styles_root.findall(".//w:style", ns):
                    style_id = style.attrib.get(f"{{{ns['w']}}}styleId", "")
                    name_node = style.find("w:name", ns)
                    if style_id and name_node is not None:
                        style_name = name_node.attrib.get(f"{{{ns['w']}}}val", "")
                        if style_name:
                            style_to_name[style_id] = style_name
            except Exception:
                style_to_name = {}

    section_title = "Policy Document"
    sections: list[tuple[str, str]] = []
    current_lines: list[str] = []

    for paragraph in doc_root.findall(".//w:body/w:p", ns):
        texts = []
        for text_node in paragraph.findall(".//w:t", ns):
            if text_node.text:
                texts.append(text_node.text)
        text = _clean_text("".join(texts))
        if not text:
            continue

        style_id = ""
        style_node = paragraph.find("w:pPr/w:pStyle", ns)
        if style_node is not None:
            style_id = style_node.attrib.get(f"{{{ns['w']}}}val", "")
        style_name = style_to_name.get(style_id, style_id)
        is_heading = style_name.lower().startswith("heading")

        if is_heading:
            if current_lines:
                sections.append((section_title, "\n".join(current_lines).strip()))
                current_lines = []
            section_title = text
            continue

        current_lines.append(text)

    if current_lines:
        sections.append((section_title, "\n".join(current_lines).strip()))

    if not sections:
        raise ValueError("Could not extract readable text from DOCX")
    return sections


def _chunk_section_text(
    section_text: str,
    *,
    section_title: str,
    start_index: int,
    chunk_size_chars: int,
    chunk_overlap_chars: int,
    min_chunk_chars: int,
) -> list[PolicyDocChunk]:
    if not section_text:
        return []

    chunks: list[PolicyDocChunk] = []
    text_len = len(section_text)
    cursor = 0
    local_index = 0
    overlap = max(0, min(chunk_overlap_chars, max(0, chunk_size_chars - 50)))

    while cursor < text_len:
        end = min(text_len, cursor + chunk_size_chars)
        if end < text_len:
            split = section_text.rfind(". ", cursor, end)
            if split > cursor + min_chunk_chars:
                end = split + 1
            else:
                split = section_text.rfind(" ", cursor, end)
                if split > cursor + min_chunk_chars:
                    end = split

        content = section_text[cursor:end].strip()
        if len(content) >= min_chunk_chars or (end >= text_len and content):
            chunks.append(
                PolicyDocChunk(
                    chunk_index=start_index + local_index,
                    content=content,
                    section_title=section_title,
                    char_start=cursor,
                    char_end=end,
                    metadata={"strategy": "heading_aware_char_chunk"},
                )
            )
            local_index += 1

        if end >= text_len:
            break
        cursor = max(cursor + 1, end - overlap)

    return chunks
